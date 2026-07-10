import openpyxl

from app.core.domain.ports.excel_parser import ExcelParser

HEADER_MAP = {
    "EORI Number": "eori_number",
    "Declarant Legal Name": "declarant_legal_name",
    "Declarant Address": "declarant_address",
    "Contact Person": "contact_person",
    "Competent Authority": "competent_authority",
    "CBAM Account Number": "cbam_account_number",
    "Data Owner": "data_owner",
    "TARIC Code": "taric_code",
    "CN Code": "cn_code",
    "Goods Description": "goods_description",
    "Sector Category": "sector_category",
    "Product Type": "product_type",
    "Import Volume": "import_volume",
    "Date of importation": "date_of_importation",
    "Country of Origin": "country_of_origin",
    "Customs Declaration Ref": "customs_declaration_ref",
    "Supplier Name": "supplier_name",
    "Notes / Comments": "notes_comments",
}


class OpenpyxlParser(ExcelParser):
    async def parse(self, file_path: str) -> list[dict]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["Template"]

        headers = [cell.value for cell in ws[1]]
        mapped_headers = [HEADER_MAP[h] for h in headers]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cleaned = {
                k: v if v is not None else ""
                for k, v in zip(mapped_headers, row)
            }
            rows.append(cleaned)

        wb.close()
        return rows
